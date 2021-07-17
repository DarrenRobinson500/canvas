from django.shortcuts import render, redirect
from .models import Db, Connection, File
from .forms import BaseForm, FileForm, RiskForm, ControlForm, ControlInstForm, ControlTestForm, ActionForm, \
    OwnerLink, BusObjLink, ProcessLink, RiskLink, ControlLink, MetricLink, IssueLink
import openpyxl as xl
from .constants import (X_SPACE, new_temp, list_temp, upload_temp, Y_SPACE, home_temp, home_link,
    home_comment, home_header, current_y, full_name, full_name_p, ind_template, canvas_temp, template_colour)

X_SPACE = 160
current_y = 0

def home(request):
    return redirect("/list/role")


def set_loc_layer(type):
    global current_y
    for level in range(1, 7):
        this_row = Db.objects.filter(type=type, level=level, visible=True)
        if this_row.count() > 0:
            current_x = int(
                20 + X_SPACE * 3 - (min(this_row.count(), 7) - 1) / 2 * X_SPACE
            )
            for item in this_row:
                item.y = current_y
                item.x = current_x
                print(type,level, this_row, item, item.y, item.x)

                item.save()
                current_x += X_SPACE
                if current_x >= X_SPACE * 7:
                    current_x = int(20)
                    current_y += Y_SPACE / 2

            current_y += Y_SPACE


def set_locations():
    global current_y
    current_y = 50
    set_loc_layer("role")
    set_loc_layer("busobj")
    set_loc_layer("obligation")
    set_loc_layer("theme")
    set_loc_layer("process")
    set_loc_layer("risk")
    set_loc_layer("metric")
    set_loc_layer("control")
    set_loc_layer("issue")


def add_connections(parent, children):
    for child in children:
        new = Connection(a=child, b=parent, visible=True)
        new.save()

def ind(request, type, id, map = False):
    item = Db.objects.get(pk=id)
    Connection.objects.all().delete()

    messages=[]
    nodes = []
    connections = []

    template = ind_template[type] + ".html"
    objects = item.links.all()
    owners = objects.filter(type="role")
    busobjs = objects.filter(type="busobj")
    obligations = objects.filter(type="obligation")
    themes = objects.filter(type="theme")
    processes = objects.filter(type="process")
    risks = objects.filter(type="risks")
    controls = objects.filter(type="control")
    controls_insts = objects.filter(type="control_inst")
    controls_tests = objects.filter(type="control_test")
    metrics = objects.filter(type="metric")
    issues = objects.filter(type="issue")
    actions = objects.filter(type="action")

    form = BusObjLink(request.POST or None, instance=item)

    if type == "role":
        if form.is_valid():
            to_remove = item.links.all().filter(type="busobj")
            for i in to_remove:
                item.links.remove(i)
            data = form.cleaned_data.get("links")
            for x in data:
                item.links.add(x)

    if type == "role" and map:
        db = Db.objects.all()
        db.update(visible=False)
        for x in busobjs:
            new = x.links.all().filter(type="busob")
            busobjs = busobjs | new
            add_connections(x,new)
            new = x.links.all().filter(type="process")
            processes = processes | new
            add_connections(x,new)
            new = x.links.all().filter(type="risk")
            risks = risks | new
            add_connections(x,new)
        for x in processes:
            processes = processes | x.links.all().filter(type="process")
            add_connections(x, processes)
            risks = risks | x.links.all().filter(type="risk")
            add_connections(x, risks)
        for x in risks:
            new = x.links.all().filter(type="risk")
            risks = risks | new
            add_connections(x, new)
            new = x.links.all().filter(type="control")
            controls = controls | new
            add_connections(x, new)

        busobjs = busobjs.distinct()
        obligations = obligations.distinct()
        themes = themes.distinct()
        processes = processes.distinct()
        risks = risks.distinct()
        metrics = metrics.distinct()
        controls = controls.distinct()
        issues = issues.distinct()

        nodes = busobjs | obligations | themes | processes | risks | controls | metrics | issues

        for node in nodes:
            node.visible = True
            node.save()
        connections = Connection.objects.all()
        set_locations()
        nodes = busobjs | obligations | themes | processes | risks | controls | metrics | issues

        if busobjs.count() == 0:
            messages.append("You currently have no business objectives. Consider adding some.")
        for busobj in busobjs:
            item_processes = busobj.links.all().filter(type="process").count()
            if item_processes == 0: messages.append(f"Your business objective '{busobj}' has no processes. Consider adding some.")

    return render(request, template, {'item': item, 'type': type, 'owners': owners,'form':form,
                                      'busobjs':busobjs, 'nodes': nodes, 'connections': connections, 'processes':processes,'risks':risks,
                                      'controls': controls, 'control_insts': controls_insts, 'control_tests': controls_tests,
                                      'actions': actions, 'messages':messages})

def update(request, type, id):
    item = Db.objects.get(pk=id)
    form = BaseForm(request.POST or None, instance=item)
    if type == "risk": form = RiskForm(request.POST or None, instance=item)
    if type == "control": form = ControlForm(request.POST or None, instance=item)
    if type == "control_inst": form = ControlInstForm(request.POST or None, instance=item)
    if type == "control_test": form = ControlTestForm(request.POST or None, instance=item)

    if form.is_valid():
        form.save()
        return redirect('/list/' + type)
    header = full_name[type] + ": " + item.name
    return render(request, 'update.html', {'header': header,'item': item, 'form': form, 'type': type})

def links(request, type, id):
    item = Db.objects.get(pk=id)
    header = full_name[item.type] + ": " + item.name
    type_name = "Select related " + full_name_p[type] + ":"
    if type == "role":
        type_name = "Owner:"

    if type == "role": form = OwnerLink(request.POST or None, instance=item)
    if type == "busobj": form = BusObjLink(request.POST or None, instance=item)
    if type == "process": form = ProcessLink(request.POST or None, instance=item)
    if type == "risk": form = RiskLink(request.POST or None, instance=item)
    if type == "control": form = ControlLink(request.POST or None, instance=item)
    if type == "metric": form = MetricLink(request.POST or None, instance=item)
    if type == "metric": form = IssueLink(request.POST or None, instance=item)

    if form.is_valid():
        to_remove = item.links.all().filter(type=type)
        for i in to_remove:
            item.links.remove(i)
        data = form.cleaned_data.get("links")
        for x in data:
            item.links.add(x)
        return redirect('/list/' + item.type)

    return render(request, 'links.html', {'header': header,'item': item, 'form': form, 'type_name': type_name})

def list(request, type):
    header = full_name_p[type]
    new_link = type + "_new"
    list = Db.objects.all().order_by("level").order_by("type")
    colour = template_colour[type][0]
    if type != "all":
        list = list.filter(type=type).order_by("level")
    return render(
        request, list_temp,
        {"header": header, "type": type, "new_link": new_link, "list": list, "colour": colour})

def new(request, type, id = None):
    if id is None:
        header = "New " + full_name[type]
    else:
        item = Db.objects.filter(id=id)[0]
        header = item.name + ": New " + full_name[type]
    form = BaseForm(request.POST)
    if type == "risk":         form = RiskForm(request.POST)
    if type == "control":      form = ControlForm(request.POST)
    if type == "control_inst": form = ControlInstForm(request.POST)
    if type == "control_test": form = ControlTestForm(request.POST)
    if type == "action":       form = ActionForm(request.POST)

    if request.method == "POST":
        if form.is_valid():
            new = form.save()
            new.type = type
            if id is not None:
                item.links.add(new)
            if new.type == "control_inst":
                new.name = "Control run"
                if new.date_performed is not None:
                    new.name = "Control run: " + str(new.date_performed)
            if new.type == "control_test":
                new.name = "Control test"
                if new.date_performed is not None:
                    new.name = "Control test: " + str(new.date_performed)
            form.save()
            if id is not None:
                path = "/ind/" + item.type + "/" + str(item.id)
                return redirect(path)
            else:
                return redirect("/list/" + type)
    return render(request, new_temp, {"header": header, "form": form})

def delete(request, type, id):
    item = Db.objects.get(pk=id)
    item.delete()
    return redirect("/list/" + type)


def upload(request):
    if request.method == "POST":
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            new_file = form.save()
            if new_file.name.find("Report") != -1:
                new_file.type = "Report Template"
            form.save()
            return redirect("files")
    else:
        form = FileForm()
    return render(request, upload_temp, {"form": form})


def load(request, type, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    messages = []
    try:
        wb = xl.load_workbook(path)
    except:
        header = full_name_p[type]
        new_link = type + "_new"
        list = Db.objects.filter(type=type)
        message = f"Could not find {file} in {path}"
        messages.append(message)
        return render(
            request,
            list_temp,
            {"header": header, "type": type, "new_link": new_link, "list": list},
        )

    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        role = sheet.cell(row, 2).value
        parent_role = sheet.cell(row, 3).value
        level = 1

        if Db.objects.filter(name=parent_role).count() > 0:
            parent_object = Db.objects.filter(name=parent_role)[0]
            level = parent_object.level + 1
        else:
            parent_object = None
            messages.append(f"No parent found for {name}")

        if Db.objects.filter(name=role).count() == 0 and role is not None:
            new_role = Db(name=role, incumbant=name, level=level, type=type)
            new_role.save()
            if parent_object is not None:
                new_role.links.add(parent_object)
            new_role.save()

    return redirect("/list/" + type)


def files(request):
    list = File.objects.all().order_by("type")
    return render(request, "files.html", {"list": list})


def file_delete(request, id):
    if request.method == "POST":
        file = File.objects.get(pk=id)
        file.delete()
    return redirect("files")


# TODO Show details of selected nodes
# TODO Add a join button
# TODO Allow join selection (to edit strength or delete)
# TODO change size to fit on page
# TODO allow user to start from one node
